"""
1. Чтение файлов с выводом команды "show tech-support"
2. Формирование структуры (structure) вида:
    [
        {
            'hostname': '...',
            'interfaces':
            {
                'ifname': '...',
                'ifaddr': ipaddress.IPv4Interface()
            },
            {
                'ifname': '...',
                'ifaddr': None
            },
            ...
        },
        {
        },
        ...
    ]
3. Составление списка сетей (net_list) с сортировкой по маске/сети
4. Запись в файл 'AddrPlan.xlsx' адресного плана в 2 стобца в worksheeet 'AddrPlan'
5. Запись в файл 'AddrPlan.xlsx' информации об интерфейсах в worksheet 'DevInfo'
6. Вывод на экран списка сетей/масок в 2 столбца
"""

import glob, ipaddress, re
from openpyxl import Workbook, load_workbook


addr_plan_file = 'AddrPlan.xlsx'

def classifier(line):
    ip_addr_search = re.search(r'ip address ((?:\d{1,3}\.){3}\d{1,3}) ((?:\d{1,3}\.){3}\d{1,3})', line)
    if ip_addr_search:
        return {'ip': ipaddress.IPv4Interface(ip_addr_search.group(1) + '/' + ip_addr_search.group(2))}
    ip_dhcp_search = re.search(r'ip address dhcp', line)
    if ip_dhcp_search:
        return {'ip': 'dhcp'}
    interface_search = re.search(r'^interface ([^ \n]+)', line.lstrip())
    if interface_search:
        return {'int': interface_search.group(1)}
    hostname_search = re.search(r'hostname ([^ \n\']+)', line.lstrip())
    if hostname_search:
        return {'host': hostname_search.group(1)}
    return {}


def struct_func(file):
    """
    Создание структуры данных с информацией об устройствах
    """
    with open(file) as openfile:
        if_list = []
        ifname = ''
        for line in openfile:
            class_dict = classifier(line)
            if 'host' in class_dict:
                hostname = class_dict['host']
            elif 'int' in class_dict:
                if ifname:
                    if_list.append(if_dict)
                ifname = class_dict['int']
                if_dict = {'ifname': ifname, 'ifaddr': None}
            elif 'ip' in class_dict:
                if_dict['ifaddr'] = class_dict['ip']

                if class_dict['ip'] == 'dhcp': continue
                netaddr = ipaddress.IPv4Network(class_dict['ip'], strict=False)
                if netaddr not in net_list:
                    net_list.append(netaddr)
        if ifname:
            if_list.append(if_dict)

        dev_dict = {'hostname': hostname,
                    'interfaces': if_list}
        structure.append(dev_dict)


def sortfunc(net):
    """
    Функция сортировки списка сетей по ключу
    """
    return (net.prefixlen << 32) + int(net.network_address)


def write_addrplan_to_file(file):
    """
    Запись в файл xlsx списка сетей
    """
    if glob.glob(file):
        wb = load_workbook(file)
        try:
            wb.remove(wb['AddrPlan'])
        except KeyError:
            pass
    else:
        wb = Workbook()

    ws = wb.create_sheet('AddrPlan')
    ws['A1'] = 'Network'
    ws['B1'] = 'Netmask'

    for net in net_list:
        ws.append({'A': str(net.network_address), 'B': str(net.netmask)})

    wb.save(file)
    wb.close()


def write_devinfo_to_file(file):
    """
    Запись в файл xlsx информации об интерфейсах устройств
    """
    if glob.glob(file):
        wb = load_workbook(file)
        try:
            wb.remove(wb['DevInfo'])
        except KeyError:
            pass
    else:
        wb = Workbook()

    ws = wb.create_sheet('DevInfo')
    ws['A1'] = 'Hostname'
    ws['B1'] = 'Interface'
    ws['C1'] = 'IPaddress'
    ws['D1'] = 'Netmask'

    for device in structure:
        for interface in device['interfaces']:
            if type(interface['ifaddr']) is ipaddress.IPv4Interface:
                C = str(interface['ifaddr'].ip)
                D = str(interface['ifaddr'].netmask)
            elif interface['ifaddr'] is None:
                C = 'None'
                D = 'None'
            elif interface['ifaddr'] == 'dhcp':
                C = 'dhcp'
                D = 'dhcp'
            else:
                print('WTF?!')

            ws.append({'A': device['hostname'],
                       'B': interface['ifname'],
                       'C': C,
                       'D': D})

    wb.save(file)
    wb.close()


def print_addrplan_to_screen():
    """
    Вывод на экран адресного плана
    """
    print('{:16s}{:16s}'.format('Network', 'Netmask'))
    print('-' * 15, '-' * 15)
    for net in net_list:
        print('{:16s}{:16s}'.format(str(net.network_address), str(net.netmask)))


structure = []
net_list = []

for file in glob.glob('../configs/*.txt'):
    struct_func(file)

net_list.sort(key=sortfunc)

write_addrplan_to_file(addr_plan_file)
write_devinfo_to_file(addr_plan_file)
print_addrplan_to_screen()
